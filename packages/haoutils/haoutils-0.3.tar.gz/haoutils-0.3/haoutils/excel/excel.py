# -*- encoding: utf-8 -*-
'''
@File    :   excel.py
@Time    :   2019/08/05 12:38:21
@Author  :   hao qihan 
@Version :   0.1
@Contact :   2263310007@qq.com
'''
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class Excel(object):
    def __init__(self,filename):
        self.filename = filename
        

    # 设置字体大小和样式
    def font(self, name="微软雅黑", size=14, bold=False, color="000000"):
        return Font(name=name, size=size, bold=bold, color=color)

    # 设置背景颜色
    def fill(self, fgColor="DDDDDD"):
        return PatternFill("solid", fgColor=fgColor)

    # 设置对其格式
    def alignment(self, format="center"):
        return Alignment(horizontal=format, vertical=format)

    # 设置边框线
    def border(self, style="thin", color="000000"):
        side = Side(style=style, color=color)
        return Border(left=side, right=side, top=side, bottom=side)

    # 添加workbook
    def create_workbook(self):
        isE = os.path.exists(self.filename)
        if isE:
            self.workbook = load_workbook(self.filename)
        else:
            self.workbook = Workbook()

    # 添加sheet
    def create_sheet(self,sheetname):
        sheet = self.workbook.create_sheet(sheetname)
        return sheet

    # 设置样式
    def set_default_style(self, cell_obj):
        cell_obj.fill = self.fill()
        cell_obj.font = self.font()
        cell_obj.alignment = self.alignment()
        cell_obj.border = self.border()

    # 向excel表添加内容
    def create_demo(self, sheet):
        for i in "ABCDEFG":
            sheet.column_dimensions[i].width = 20
        row = ["运营商", "ismi", "手机型号", "mcc", "mnc", "采集日期"]
        for index, i in enumerate(row):
            xx = sheet.cell(1,index+1,index)
            self.set_style(xx)
        
    def save(self):
        self.workbook.save(self.filename)




