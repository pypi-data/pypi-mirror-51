# -*- coding: UTF-8 -*-
#! python3

"""
    Get metadatas from Isogeo and store it into a Excel worksheet. 

"""

# ##############################################################################
# ########## Libraries #############
# ##################################

# Standard library
import logging
from collections import Counter
from collections.abc import KeysView
from pathlib import Path
from urllib.parse import urlparse

# 3rd party library
from isogeo_pysdk import IsogeoTranslator, IsogeoUtils, Metadata, Share
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# custom submodules
from isogeopyxl.matrix import (
    RASTER_COLUMNS,
    RESOURCE_COLUMNS,
    SERVICE_COLUMNS,
    VECTOR_COLUMNS,
)
from isogeopyxl.utils import Formatter, Stats

# ##############################################################################
# ############ Globals ############
# #################################


logger = logging.getLogger("isogeo2office")
utils = IsogeoUtils()

# ##############################################################################
# ########## Classes ###############
# ##################################


class Isogeo2xlsx(Workbook):
    """Used to store Isogeo API results into an Excel worksheet (.xlsx)

    :param str lang: selected language for output
    :param str url_base_edit: base url to format edit links (basically app.isogeo.com)
    :param str url_base_view: base url to format view links (basically open.isogeo.com)
    """

    cols_v = [
        "Titre",  # A
        "Nom",  # B
        "Résumé",  # C
        "Emplacement",  # D
        "Groupe de travail",  # E
        "Mots-clés",  # F
        "Thématique(s) INSPIRE",  # G
        "Conformité INSPIRE",  # H
        "Contexte de collecte",  # I
        "Méthode de collecte",  # J
        "Début de validité",  # K
        "Fin de validité",  # L
        "Fréquence de mise à jour",  # M
        "Commentaire",  # N
        "Création",  # O
        "# mises à jour",  # P
        "Dernière mise à jour",  # Q
        "Publication",  # R
        "Format (version - encodage)",  # S
        "SRS (EPSG)",  # T
        "Emprise",  # U
        "Géométrie",  # V
        "Résolution",  # W
        "Echelle",  # X
        "# Objets",  # Y
        "# Attributs",  # Z
        "Attributs (A-Z)",  # AA
        "Spécifications",  # AB
        "Cohérence topologique",  # AC
        "Conditions",  # AD
        "Limitations",  # AE
        "Contacts",  # AF
        "Téléchargeable",  # AG
        "Visualisable",  # AH
        "Autres",  # AI
        "Editer",  # AJ
        "Consulter",  # AK
        "MD - ID",  # AL
        "MD - Création",  # AM
        "MD - Modification",  # AN
        "MD - Langue",  # AO
    ]

    cols_r = [
        "Titre",  # A
        "Nom",  # B
        "Résumé",  # C
        "Emplacement",  # D
        "Groupe de travail",  # E
        "Mots-clés",  # F
        "Thématique(s) INSPIRE",  # G
        "Conformité INSPIRE",  # H
        "Contexte de collecte",  # I
        "Méthode de collecte",  # J
        "Début de validité",  # K
        "Fin de validité",  # L
        "Fréquence de mise à jour",  # M
        "Commentaire",  # N
        "Création",  # O
        "# mises à jour",  # P
        "Dernière mise à jour",  # Q
        "Publication",  # R
        "Format (version - encodage)",  # S
        "SRS (EPSG)",  # T
        "Emprise",  # U
        "Résolution",  # V
        "Echelle",  # W
        "Spécifications",  # X
        "Cohérence topologique",  # Y
        "Conditions",  # Z
        "Limitations",  # AA
        "Contacts",  # AB
        "Téléchargeable",  # AC
        "Visualisable",  # AD
        "Autres",  # AE
        "Editer",  # AF
        "Consulter",  # AG
        "MD - ID",  # AH
        "MD - Création",  # AI
        "MD - Modification",  # AJ
        "MD - Langue",  # AK
    ]

    cols_s = [
        "Titre",  # A
        "Nom",  # B
        "Résumé",  # C
        "Emplacement",  # D
        "Groupe de travail",  # E
        "Mots-clés",  # F
        "Conformité INSPIRE",  # G
        "Création",  # H
        "# mises à jour",  # I
        "Dernière mise à jour",  # J
        "Publication",  # K
        "Format (version)",  # L
        "Emprise",  # M
        "Spécifications",  # N
        "Conditions",  # O
        "Limitations",  # P
        "Contacts",  # Q
        "Téléchargeable",  # R
        "Visualisable",  # S
        "Autres",  # T
        "Editer",  # U
        "Consulter",  # V
        "MD - ID",  # W
        "MD - Création",  # X
        "MD - Modification",  # Y
        "MD - Langue",  # Z
    ]

    cols_rz = [
        "Titre",  # A
        "Résumé",  # B
        "Emplacement",  # C
        "Groupe de travail",  # D
        "Mots-clés",  # E
        "Création",  # F
        "# mises à jour",  # G
        "Dernière mise à jour",  # H
        "Publication",  # I
        "Format (version)",  # J
        "Conditions",  # K
        "Limitations",  # L
        "Contacts",  # M
        "Téléchargeable",  # N
        "Visualisable",  # O
        "Autres",  # P
        "Editer",  # Q
        "Consulter",  # R
        "MD - ID",  # S
        "MD - Création",  # T
        "MD - Modification",  # U
        "MD - Langue",  # V
    ]

    cols_fa = ["Nom", "Occurrences"]  # A  # B

    def __init__(
        self, lang: str = "FR", url_base_edit: str = "", url_base_view: str = ""
    ):
        """Instanciating the output workbook.

        :param str lang: selected language for output
        :param str url_base_edit: base url to format edit links (basically app.isogeo.com)
        :param str url_base_view: base url to format view links (basically open.isogeo.com)
        """
        super(Isogeo2xlsx, self).__init__()
        # super(Isogeo2xlsx, self).__init__(write_only=True)

        self.stats = Stats()

        # URLS
        utils.app_url = url_base_edit  # APP
        utils.oc_url = url_base_view  # OpenCatalog url

        # styles
        s_date = NamedStyle(name="date")
        s_wrap = NamedStyle(name="wrap")
        s_wrap.alignment = Alignment(wrap_text=True)
        self.add_named_style(s_date)
        self.add_named_style(s_wrap)

        # deleting the default worksheet
        ws = self.active
        self.remove(ws)

        # LOCALE
        if lang.lower() == "fr":
            s_date.number_format = "dd/mm/yyyy"
            self.dates_fmt = "DD/MM/YYYY"
            self.locale_fmt = "fr_FR"
        else:
            s_date.number_format = "yyyy/mm/dd"
            self.dates_fmt = "YYYY/MM/DD"
            self.locale_fmt = "uk_UK"

        # TRANSLATIONS
        self.tr = IsogeoTranslator(lang).tr

        # FORMATTER
        self.fmt = Formatter(output_type="Excel")

    # ------------ Setting workbook ---------------------

    def set_worksheets(
        self,
        auto: KeysView = None,
        vector: bool = 1,
        raster: bool = 1,
        service: bool = 1,
        resource: bool = 1,
        dashboard: bool = 0,
        attributes: bool = 0,
        fillfull: bool = 0,
        inspire: bool = 0,
    ):
        """Adds new sheets depending on present metadata types in isogeo API
        search tags.

        :param list auto: typically auto=search_results.get('tags').keys()
        :param bool vector: add vector sheet
        :param bool raster: add raster sheet
        :param bool service: add service sheet
        :param bool resource: add resource sheet
        :param bool dashboard: add dashboard sheet
        :param bool attributes: add attributes sheet - only if vector is True too
        :param bool fillfull: add fillfull sheet
        :param bool inspire: add inspire sheet
        """
        if isinstance(auto, KeysView):
            logger.info("Automatic sheets creation based on tags")
            if "type:vector-dataset" in auto:
                vector = 1
            else:
                vector = 0
            if "type:raster-dataset" in auto:
                raster = 1
            else:
                raster = 0
                pass
            if "type:resource" in auto:
                resource = 1
            else:
                resource = 0
                pass
            if "type:service" in auto:
                service = 1
            else:
                service = 0
                pass
        else:
            raise TypeError(
                "'auto' must be a KeysView (dict.keys()),"
                " from Isogeo search request, not {}".format(type(auto))
            )

        # SHEETS & HEADERS
        if dashboard:
            self.ws_d = self.create_sheet(title="Tableau de bord")
            # headers
            # self.ws_f.append([i for i in self.cols_v])
            # styling
            # for i in self.cols_v:
            #     self.ws_v.cell(row=1,
            #                    column=self.cols_v.index(i) + 1).style = "Headline 2"
            # initialize line counter
            self.idx_d = 1
            # log
            logger.info("Dashboard sheet added")
        else:
            pass
        if fillfull:
            self.ws_f = self.create_sheet(title="Progression catalogage")
            # headers
            # self.ws_f.append([i for i in self.cols_v])
            # styling
            # for i in self.cols_v:
            #     self.ws_v.cell(row=1,
            #                    column=self.cols_v.index(i) + 1).style = "Headline 2"
            # initialize line counter
            self.idx_f = 1
            # log
            logger.info("Fillfull sheet added")
        else:
            pass
        if inspire:
            self.ws_i = self.create_sheet(title="Directive INSPIRE")
            # headers
            # self.ws_f.append([i for i in self.cols_v])
            # styling
            # for i in self.cols_v:
            #     self.ws_v.cell(row=1,
            #                    column=self.cols_v.index(i) + 1).style = "Headline 2"
            # initialize line counter
            self.idx_i = 1
            # log
            logger.info("INSPIRE sheet added")
        else:
            pass
        if vector:
            self.ws_v = self.create_sheet(title="Vecteurs")
            # headers
            self.ws_v.append([i for i in self.cols_v])
            # styling
            for i in self.cols_v:
                self.ws_v.cell(
                    row=1, column=self.cols_v.index(i) + 1
                ).style = "Headline 2"
            # initialize line counte
            self.idx_v = 1
            # log
            logger.info("Vectors sheet added")
            # feature attributes analisis
            if attributes:
                self.ws_fa = self.create_sheet(title="Attributs")
                self.idx_fa = 1
                self.fa_all = []
                # headers
                self.ws_fa.append([i for i in self.cols_fa])
                # styling
                for i in self.cols_fa:
                    self.ws_fa.cell(
                        row=1, column=self.cols_fa.index(i) + 1
                    ).style = "Headline 2"
                logger.info("Feature attributes sheet added")
            else:
                pass
        else:
            pass

        if raster:
            self.ws_r = self.create_sheet(title="Raster")
            # headers
            self.ws_r.append([i for i in self.cols_r])
            # styling
            for i in self.cols_r:
                self.ws_r.cell(
                    row=1, column=self.cols_r.index(i) + 1
                ).style = "Headline 2"
            # initialize line counter
            self.idx_r = 1
            # log
            logger.info("Rasters sheet added")
        else:
            pass

        if service:
            self.ws_s = self.create_sheet(title="Services")
            # headers
            self.ws_s.append([i for i in self.cols_s])
            # styling
            for i in self.cols_s:
                self.ws_s.cell(
                    row=1, column=self.cols_s.index(i) + 1
                ).style = "Headline 2"
            # initialize line counter
            self.idx_s = 1
            # log
            logger.info("Services sheet added")
        else:
            pass

        if resource:
            self.ws_rz = self.create_sheet(title="Ressources")
            # headers
            self.ws_rz.append([i for i in self.cols_rz])
            # styling
            for i in self.cols_rz:
                self.ws_rz.cell(
                    row=1, column=self.cols_rz.index(i) + 1
                ).style = "Headline 2"
            # initialize line counter
            self.idx_rz = 1
            # log
            logger.info("Resources sheet added")
        else:
            pass

    # ------------ Writing metadata ---------------------
    def store_metadatas(self, metadata: Metadata, share: Share = None):
        """Write metadata into the worksheet.

        :param Metadata metadata: metadata object to write
        """
        # check input
        if not isinstance(metadata, Metadata):
            raise TypeError("Export expects a Metadata object.")
        # generic export
        self.share = share
        # store depending on metadata type
        if metadata.type == "vectorDataset":
            self.idx_v += 1
            self.store_md_generic(metadata, self.ws_v, self.idx_v)
            self.stats.md_types_repartition["vector"] += 1
            self.store_md_vector(metadata, self.ws_v, self.idx_v)
        elif metadata.type == "rasterDataset":
            self.idx_r += 1
            self.store_md_generic(metadata, self.ws_r, self.idx_r)
            self.stats.md_types_repartition["raster"] += 1
            self.store_md_raster(metadata, self.ws_r, self.idx_r)
        elif metadata.type == "service":
            self.idx_s += 1
            self.store_md_generic(metadata, self.ws_s, self.idx_s)
            self.stats.md_types_repartition["service"] += 1
            self.store_md_service(metadata, self.ws_s, self.idx_s)
        elif metadata.type == "resource":
            self.idx_rz += 1
            self.store_md_generic(metadata, self.ws_rz, self.idx_rz)
            self.stats.md_types_repartition["resource"] += 1
            self.store_md_resource(metadata, self.ws_rz, self.idx_rz)
        else:
            logger.error(
                "Type of metadata is not recognized/handled: {}".format(metadata.type)
            )

    def store_md_generic(self, md: Metadata, ws: Worksheet, idx: int):
        """Exports genreic metadata attributes into Excel worksheet with some dynamic
        adaptations based on metadata type.

        :param Metadata md: metadata object to export
        :param Worksheet ws: Excel worksheet to store the exported info
        :param int idx: row index in the worksheet
        """
        if md.type == "rasterDataset":
            colsref = RASTER_COLUMNS
        elif md.type == "resource":
            colsref = RESOURCE_COLUMNS
        elif md.type == "service":
            colsref = SERVICE_COLUMNS
        elif md.type == "vectorDataset":
            colsref = VECTOR_COLUMNS
        else:
            raise TypeError("Unknown metadata type: {}".format(md.type))

        logger.debug(
            "Start storing metadata {} ({}) using the matching reference columns for type of {} ...".format(
                md.title_or_name(slugged=1), md._id, md.type
            )
        )

        # -- IDENTIFICATION ------------------------------------------------------------
        if md.title:
            ws["{}{}".format(colsref.get("title"), idx)] = md.title
        if md.name:
            ws["{}{}".format(colsref.get("name"), idx)] = md.name
        if md.abstract:
            ws["{}{}".format(colsref.get("abstract"), idx)] = md.abstract

        # path to source
        try:
            src_path = Path(str(md.path))
        except OSError as e:
            logger.debug(
                "Metadata.path value is not a valid system path. Maybe an URL? Original error: {}".format(
                    e
                )
            )
            print(urlparse(md.path).scheme)
            urlparse(md.path).scheme != ""

        if isinstance(md.path, Path) and md.type != "service":
            if src_path.is_file():
                link_path = r'=HYPERLINK("{0}","{1}")'.format(
                    src_path.parent, src_path.resolve()
                )
                ws["{}{}".format(colsref.get("path"), idx)] = link_path
                logger.debug("Path reachable: {}".format(src_path))
            else:
                ws["{}{}".format(colsref.get("path"), idx)] = str(src_path.resolve())
                logger.debug(
                    "Path not recognized nor reachable: {}".format(str(src_path))
                )
        elif md.path and md.type == "service":
            link_path = r'=HYPERLINK("{0}","{1}")'.format(md.path, md.path)
            ws["{}{}".format(colsref.get("path"), idx)] = link_path
        else:
            pass

        # -- TAGS ----------------------------------------------------------------------
        keywords = []
        inspire = []
        if md.keywords:
            for k in md.keywords:
                if k.get("_tag").startswith("keyword:is"):
                    keywords.append(k.get("text"))
                elif k.get("_tag").startswith("keyword:in"):
                    inspire.append(k.get("text"))
                else:
                    logger.info("Unknown keyword type: " + k.get("_tag"))
                    continue
            if keywords:
                ws["{}{}".format(colsref.get("keywords"), idx)] = " ;\n".join(
                    sorted(keywords)
                )
            if inspire:
                ws["{}{}".format(colsref.get("inspireThemes"), idx)] = " ;\n".join(
                    sorted(inspire)
                )
        else:
            self.stats.md_empty_fields[md._id].append("keywords")
            logger.info("Vector dataset without any keyword or INSPIRE theme")

        # INSPIRE conformity
        ws["{}{}".format(colsref.get("inspireConformance"), idx)] = (
            "conformity:inspire" in md.tags
        )

        # owner
        ws["{}{}".format(colsref.get("_creator"), idx)] = next(
            v for k, v in md.tags.items() if "owner:" in k
        )

        # -- HISTORY -------------------------------------------------------------------
        if md.collectionContext:
            ws[
                "{}{}".format(colsref.get("collectionContext"), idx)
            ] = md.collectionContext
        if md.collectionMethod:
            ws[
                "{}{}".format(colsref.get("collectionMethod"), idx)
            ] = md.collectionMethod

        # validity
        if md.validFrom:
            ws["{}{}".format(colsref.get("validFrom"), idx)] = utils.hlpr_datetimes(
                md.validFrom
            )
            ws["{}{}".format(colsref.get("validFrom"), idx)].style = "date"

        if md.validTo:
            ws["{}{}".format(colsref.get("validTo"), idx)] = utils.hlpr_datetimes(
                md.validTo
            )
            ws["{}{}".format(colsref.get("validTo"), idx)].style = "date"

        if md.updateFrequency:
            ws["{}{}".format(colsref.get("updateFrequency"), idx)] = md.updateFrequency
        if md.validityComment:
            ws["{}{}".format(colsref.get("validityComment"), idx)] = md.validityComment

        # -- EVENTS --------------------------------------------------------------------
        # data creation date
        if md.created:
            ws["{}{}".format(colsref.get("created"), idx)] = utils.hlpr_datetimes(
                md.created
            )
            ws["{}{}".format(colsref.get("created"), idx)].style = "date"

        # events count
        if md.events:
            ws["{}{}".format(colsref.get("events"), idx)] = len(md.events)

        # data last update
        if md.modified:
            ws["{}{}".format(colsref.get("modified"), idx)] = utils.hlpr_datetimes(
                md.modified
            )
            ws["{}{}".format(colsref.get("modified"), idx)].style = "date"

        # -- TECHNICAL -----------------------------------------------------------------
        # format
        if md.format and md.type in ("rasterDataset", "vectorDataset"):
            format_lbl = next(v for k, v in md.tags.items() if "format:" in k)
            ws["{}{}".format(colsref.get("format"), idx)] = "{0} ({1} - {2})".format(
                format_lbl, md.formatVersion, md.encoding
            )
        elif md.format:
            ws["{}{}".format(colsref.get("format"), idx)] = "{0} {1}".format(
                md.format, md.formatVersion
            )
        else:
            pass

        # SRS
        if isinstance(md.coordinateSystem, dict):
            ws[
                "{}{}".format(colsref.get("coordinateSystem"), idx)
            ] = "{0} ({1})".format(
                md.coordinateSystem.get("name"), md.coordinateSystem.get("code")
            )

        # bounding box (envelope)
        if md.envelope and md.envelope.get("bbox"):
            coords = md.envelope.get("coordinates")
            if md.envelope.get("type") == "Polygon":
                bbox = ",\n".join(
                    format(coord, ".4f") for coord in md.envelope.get("bbox")
                )
            elif md.envelope.get("type") == "Point":
                bbox = "Centroïde : {}{}".format(coords[0], coords[1])
            else:
                bbox = ",\n".join(
                    format(coord, ".4f") for coord in md.envelope.get("bbox")
                )
            ws["{}{}".format(colsref.get("envelope"), idx)] = bbox

        # geometry
        if md.geometry:
            ws["{}{}".format(colsref.get("geometry"), idx)] = md.geometry

        # resolution
        if md.distance:
            ws["{}{}".format(colsref.get("distance"), idx)] = md.distance

        # scale
        if md.scale:
            ws["{}{}".format(colsref.get("scale"), idx)] = md.scale

        # features objects
        if md.features:
            ws["{}{}".format(colsref.get("features"), idx)] = md.features

        # -- QUALITY -------------------------------------------------------------------
        if md.specifications:
            ws["{}{}".format(colsref.get("specifications"), idx)] = " ;\n".join(
                self.fmt.specifications(md.specifications)
            )

        # topology
        if md.topologicalConsistency:
            ws["AC{}".format(idx)] = md.topologicalConsistency

        # -- FEATURE ATTRIBUTES --------------------------------------------------------
        if md.type == "vectorDataset" and isinstance(md.featureAttributes, list):
            fields = md.featureAttributes

            # count
            ws["{}{}".format(colsref.get("featureAttributesCount"), idx)] = len(fields)
            # alphabetic list
            fields_cct = sorted(
                [
                    "{} ({}) - Type : {} - Descripion : {:.20} [...]".format(
                        field.get("name"),
                        field.get("alias"),
                        field.get("dataType"),
                        # field.get("language"),
                        field.get("description", ""),
                    )
                    for field in fields
                ]
            )
            ws["{}{}".format(colsref.get("featureAttributes"), idx)] = " ;\n".join(
                fields_cct
            )
            # if attributes analisis is activated, append fields dict
            if hasattr(self, "ws_fa"):
                self.fa_all.append(fields)
            else:
                pass

        # -- CGUs ----------------------------------------------------------------------
        if md.conditions:
            ws["{}{}".format(colsref.get("conditions"), idx)] = " ;\n".join(
                self.fmt.conditions(md.conditions)
            )

        # -- LIMITATIONS ---------------------------------------------------------------
        if md.limitations:
            ws["{}{}".format(colsref.get("limitations"), idx)] = " ;\n".join(
                self.fmt.limitations(md.limitations)
            )

        # -- CONTACTS ------------------------------------------------------------------
        if md.contacts:
            contacts = [
                "{0} ({1})".format(
                    contact.get("contact").get("name"),
                    contact.get("contact").get("email"),
                )
                for contact in md.contacts
            ]
            ws["{}{}".format(colsref.get("contacts"), idx)] = " ;\n".join(contacts)

        # -- ACTIONS -------------------------------------------------------------------
        ws["{}{}".format(colsref.get("hasLinkDownload"), idx)] = (
            "action:download" in md.tags
        )
        ws["{}{}".format(colsref.get("hasLinkView"), idx)] = "action:view" in md.tags
        ws["{}{}".format(colsref.get("hasLinkOther"), idx)] = "action:other" in md.tags

        # -- METADATA ------------------------------------------------------------------
        # id
        ws["{}{}".format(colsref.get("_id"), idx)] = md._id

        # creation
        if md._created:
            ws["{}{}".format(colsref.get("_created"), idx)] = utils.hlpr_datetimes(
                md._created
            )
            ws["{}{}".format(colsref.get("_created"), idx)].style = "date"

        # last update
        if md._modified:
            ws["{}{}".format(colsref.get("_modified"), idx)] = utils.hlpr_datetimes(
                md._modified
            )
            ws["{}{}".format(colsref.get("_modified"), idx)].style = "date"

        # edit
        # ws["{}{}".format(colsref.get("linkEdit"), idx)] = md.admin_url(self.url_base_edit) + "identification"
        ws["{}{}".format(colsref.get("linkEdit"), idx)] = utils.get_edit_url(md)
        if self.share is not None:
            link_visu = utils.get_view_url(
                md_id=md._id, share_id=self.share._id, share_token=self.share.urlToken
            )
            ws["{}{}".format(colsref.get("linkView"), idx)] = link_visu

        # lang
        ws["{}{}".format(colsref.get("language"), idx)] = md.language

    def store_md_vector(self, md: Metadata, ws: Worksheet, idx: int):
        """ TO DOCUMENT
        """

        # STYLING
        ws.row_dimensions[idx].height = 35  # line height - see #52
        ws["C{}".format(idx)].style = "wrap"
        ws["F{}".format(idx)].style = "wrap"
        ws["G{}".format(idx)].style = "wrap"
        ws["I{}".format(idx)].style = "wrap"
        ws["J{}".format(idx)].style = "wrap"
        ws["K{}".format(idx)].style = "date"
        ws["L{}".format(idx)].style = "date"
        ws["U{}".format(idx)].style = "wrap"
        ws["AA{}".format(idx)].style = "wrap"
        ws["AB{}".format(idx)].style = "wrap"
        ws["AC{}".format(idx)].style = "wrap"
        ws["AD{}".format(idx)].style = "wrap"
        ws["AE{}".format(idx)].style = "wrap"
        ws["AG{}".format(idx)].style = "wrap"
        ws["AH{}".format(idx)].style = "wrap"

        # LOG
        logger.info(
            "Vector metadata stored: {} ({})".format(
                md.title_or_name(slugged=1), md._id
            )
        )

    def store_md_raster(self, md: Metadata, ws: Worksheet, idx: int):
        """ TO DOCUMENT
        """

        # STYLING
        ws.row_dimensions[idx].height = 35  # line height - see #52
        ws["C{}".format(idx)].style = "wrap"
        ws["F{}".format(idx)].style = "wrap"
        ws["G{}".format(idx)].style = "wrap"
        ws["I{}".format(idx)].style = "wrap"
        ws["J{}".format(idx)].style = "wrap"
        ws["K{}".format(idx)].style = "date"
        ws["L{}".format(idx)].style = "date"
        ws["U{}".format(idx)].style = "wrap"
        ws["X{}".format(idx)].style = "wrap"
        ws["Y{}".format(idx)].style = "wrap"
        ws["Z{}".format(idx)].style = "wrap"
        ws["AA{}".format(idx)].style = "wrap"
        ws["AC{}".format(idx)].style = "wrap"
        ws["AD{}".format(idx)].style = "wrap"

        # LOG
        logger.info("Raster metadata stored: {} ({})".format(md.name, md._id))

        # end of method
        return

    def store_md_service(self, md: Metadata, ws: Worksheet, idx: int):
        """ TO DOCUMENT
        """

        # STYLING
        ws.row_dimensions[idx].height = 35  # line height - see #52
        ws["C{}".format(idx)].style = "wrap"
        ws["F{}".format(idx)].style = "wrap"
        ws["M{}".format(idx)].style = "wrap"
        ws["N{}".format(idx)].style = "wrap"
        ws["O{}".format(idx)].style = "wrap"
        ws["P{}".format(idx)].style = "wrap"
        ws["R{}".format(idx)].style = "wrap"
        ws["S{}".format(idx)].style = "wrap"

        # LOG
        logger.info("Service metadata stored: {} ({})".format(md.name, md._id))

        # end of method
        return

    def store_md_resource(self, md: Metadata, ws: Worksheet, idx: int):
        """ TO DOCUMENT
        """

        # STYLING
        ws.row_dimensions[idx].height = 35  # line height - see #52
        ws["C{}".format(idx)].style = "wrap"
        ws["F{}".format(idx)].style = "wrap"
        ws["M{}".format(idx)].style = "wrap"
        ws["N{}".format(idx)].style = "wrap"
        ws["O{}".format(idx)].style = "wrap"
        ws["P{}".format(idx)].style = "wrap"
        ws["R{}".format(idx)].style = "wrap"
        ws["S{}".format(idx)].style = "wrap"

        # LOG
        logger.info("Resource metadata stored: {} ({})".format(md.name, md._id))

        # end of method
        return

    # ------------ Analisis --------------------------------------------------
    def analisis_attributes(self):
        """Perform feature attributes analisis and write results into the
        dedicatedWworksheet."""
        # local arrays
        fa_names = []
        fa_types = []
        fa_alias = []
        fa_descr = []

        # parsing
        for dico_fa in self.fa_all:
            for fa in dico_fa:
                fa_names.append(fa.get("name"))
                # fa_alias.append(fa.get("alias", "NR"))
                # fa_types.append(fa.get("dataType"))
                # fa_descr.append(fa.get("description", "NR"))
                del fa

        # stats
        frq_names = Counter(fa_names)
        frq_alias = Counter(fa_alias)
        frq_types = Counter(fa_types)
        frq_descr = Counter(fa_descr)

        # write
        ws = self.ws_fa
        for fa in frq_names:
            self.idx_fa += 1
            ws["A{}".format(self.idx_fa)] = fa
            ws["B{}".format(self.idx_fa)] = frq_names.get(fa)

    # ------------ CustomizeWworksheet ----------------------------------------
    def tunning_worksheets(self):
        """Automate"""
        for sheet in self.worksheets:
            # Freezing panes
            c_freezed = sheet["B2"]
            sheet.freeze_panes = c_freezed

            # Print properties
            sheet.print_options.horizontalCentered = True
            sheet.print_options.verticalCentered = True
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

            # Others properties
            wsprops = sheet.sheet_properties
            wsprops.filterMode = True

            # enable filters
            sheet.auto_filter.ref = str("A1:{}{}").format(
                get_column_letter(sheet.max_column), sheet.max_row
            )


# #############################################################################
# ##### Stand alone program ########
# ##################################
if __name__ == "__main__":
    """ Standalone execution and development tests """
    wb = Isogeo2xlsx(url_base="https://open.isogeo.com")
