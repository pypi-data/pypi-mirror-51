# -*- coding: UTF-8 -*-
#! python3

"""
    Get metadatas from Isogeo and store it into a Excel worksheet. 

"""

# ##############################################################################
# ########## Libraries #############
# ##################################

# Standard library
import logging
from collections import Counter
from collections.abc import KeysView
from pathlib import Path
from urllib.parse import urlparse

# 3rd party library
from isogeo_pysdk import IsogeoUtils, Metadata, Share
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# custom submodules
from isogeotoxlsx.i18n import I18N_EN, I18N_FR
from isogeotoxlsx.matrix import (
    ColumnPattern,
    RASTER_COLUMNS,
    RESOURCE_COLUMNS,
    SERVICE_COLUMNS,
    VECTOR_COLUMNS,
)
from isogeotoxlsx.utils import Formatter, Stats

# ##############################################################################
# ############ Globals ############
# #################################

logger = logging.getLogger("isogeotoxlsx")
utils = IsogeoUtils()

# ##############################################################################
# ########## Classes ###############
# ##################################


class Isogeo2xlsx(Workbook):
    """Used to store Isogeo API results into an Excel worksheet (.xlsx)

    :param str lang: selected language for output
    :param str url_base_edit: base url to format edit links (basically app.isogeo.com)
    :param str url_base_view: base url to format view links (basically open.isogeo.com)
    """

    cols_fa = ["Nom", "Occurrences"]  # A  # B

    def __init__(
        self, lang: str = "FR", url_base_edit: str = "", url_base_view: str = ""
    ):
        """Instanciating the output workbook.

        :param str lang: selected language for output
        :param str url_base_edit: base url to format edit links (basically app.isogeo.com)
        :param str url_base_view: base url to format view links (basically open.isogeo.com)
        """
        super(Isogeo2xlsx, self).__init__()
        # super(Isogeo2xlsx, self).__init__(write_only=True)

        self.stats = Stats()

        # URLS
        utils.app_url = url_base_edit  # APP
        utils.oc_url = url_base_view  # OpenCatalog url

        # styles
        s_date = NamedStyle(name="date")
        s_wrap = NamedStyle(name="wrap")
        s_wrap.alignment = Alignment(wrap_text=True)
        self.add_named_style(s_date)
        self.add_named_style(s_wrap)

        # deleting the default worksheet
        ws = self.active
        self.remove(ws)

        # LOCALE
        if lang.lower() == "fr":
            s_date.number_format = "dd/mm/yyyy"
            self.dates_fmt = "DD/MM/YYYY"
            self.locale_fmt = "fr_FR"
            self.tr = I18N_FR
        else:
            s_date.number_format = "yyyy/mm/dd"
            self.dates_fmt = "YYYY/MM/DD"
            self.locale_fmt = "uk_UK"
            self.tr = I18N_EN

        # FORMATTER
        self.fmt = Formatter(output_type="Excel")

    # ------------ Setting workbook ---------------------

    def set_worksheets(
        self,
        auto: KeysView = None,
        vector: bool = 1,
        raster: bool = 1,
        service: bool = 1,
        resource: bool = 1,
        dashboard: bool = 0,
        attributes: bool = 0,
        fillfull: bool = 0,
        inspire: bool = 0,
    ):
        """Adds new sheets depending on present metadata types in isogeo API
        search tags.

        :param list auto: typically auto=search_results.get('tags').keys()
        :param bool vector: add vector sheet
        :param bool raster: add raster sheet
        :param bool service: add service sheet
        :param bool resource: add resource sheet
        :param bool dashboard: add dashboard sheet
        :param bool attributes: add attributes sheet - only if vector is True too
        :param bool fillfull: add fillfull sheet
        :param bool inspire: add inspire sheet
        """
        if isinstance(auto, KeysView):
            logger.info("Automatic sheets creation based on tags")
            if "type:vector-dataset" in auto:
                vector = 1
                self.columns_vector = {
                    k: ColumnPattern._make(v) for k, v in VECTOR_COLUMNS.items()
                }
            else:
                vector = 0
            if "type:raster-dataset" in auto:
                raster = 1
                self.columns_raster = {
                    k: ColumnPattern._make(v) for k, v in RASTER_COLUMNS.items()
                }
            else:
                raster = 0
                pass
            if "type:resource" in auto:
                resource = 1
                self.columns_resource = {
                    k: ColumnPattern._make(v) for k, v in RESOURCE_COLUMNS.items()
                }
            else:
                resource = 0
                pass
            if "type:service" in auto:
                service = 1
                self.columns_service = {
                    k: ColumnPattern._make(v) for k, v in SERVICE_COLUMNS.items()
                }
            else:
                service = 0
                pass
        else:
            raise TypeError(
                "'auto' must be a KeysView (dict.keys()),"
                " from Isogeo search request, not {}".format(type(auto))
            )

        # SHEETS & HEADERS
        if dashboard:
            self.ws_d = self.create_sheet(title="Tableau de bord")
            # headers
            # self.ws_f.append([i for i in self.cols_v])
            # styling
            # for i in self.cols_v:
            #     self.ws_v.cell(row=1,
            #                    column=self.cols_v.index(i) + 1).style = "Headline 2"
            # initialize line counter
            self.idx_d = 1
            # log
            logger.info("Dashboard sheet added")
        else:
            pass
        if fillfull:
            self.ws_f = self.create_sheet(title="Progression catalogage")
            # headers
            # self.ws_f.append([i for i in self.cols_v])
            # styling
            # for i in self.cols_v:
            #     self.ws_v.cell(row=1,
            #                    column=self.cols_v.index(i) + 1).style = "Headline 2"
            # initialize line counter
            self.idx_f = 1
            # log
            logger.info("Fillfull sheet added")
        else:
            pass
        if inspire:
            self.ws_i = self.create_sheet(title="Directive INSPIRE")
            # headers
            # self.ws_f.append([i for i in self.cols_v])
            # styling
            # for i in self.cols_v:
            #     self.ws_v.cell(row=1,
            #                    column=self.cols_v.index(i) + 1).style = "Headline 2"
            # initialize line counter
            self.idx_i = 1
            # log
            logger.info("INSPIRE sheet added")
        else:
            pass
        if vector:
            self.ws_v = self.create_sheet(title=self.tr.get("vector"))
            # headers
            self.headers_writer(ws=self.ws_v, columns=self.columns_vector)
            # initialize line count
            self.idx_v = 1
            # log
            logger.info("Vectors sheet added")
            # feature attributes analisis
            if attributes:
                self.ws_fa = self.create_sheet(title="Attributs")
                self.idx_fa = 1
                self.fa_all = []
                # headers
                self.ws_fa.append([i for i in self.cols_fa])
                # styling
                for i in self.cols_fa:
                    self.ws_fa.cell(
                        row=1, column=self.cols_fa.index(i) + 1
                    ).style = "Headline 2"
                logger.info("Feature attributes sheet added")
            else:
                pass
        else:
            pass

        if raster:
            self.ws_r = self.create_sheet(title=self.tr.get("raster"))
            # headers
            self.headers_writer(ws=self.ws_r, columns=self.columns_raster)
            # initialize line counter
            self.idx_r = 1
            # log
            logger.info("Rasters sheet added")
        else:
            pass

        if service:
            self.ws_s = self.create_sheet(title=self.tr.get("service"))
            # headers
            self.headers_writer(ws=self.ws_s, columns=self.columns_service)
            # initialize line counter
            self.idx_s = 1
            # log
            logger.info("Services sheet added")
        else:
            pass

        if resource:
            self.ws_rz = self.create_sheet(title=self.tr.get("resource"))
            # headers
            self.headers_writer(ws=self.ws_rz, columns=self.columns_resource)
            # initialize line counter
            self.idx_rz = 1
            # log
            logger.info("Resources sheet added")
        else:
            pass

    # ------------ Writing metadata ---------------------
    def store_metadatas(self, metadata: Metadata, share: Share = None):
        """Write metadata into the worksheet.

        :param Metadata metadata: metadata object to write
        """
        # check input
        if not isinstance(metadata, Metadata):
            raise TypeError("Export expects a Metadata object.")
        # generic export
        self.share = share
        # store depending on metadata type
        if metadata.type == "vectorDataset":
            self.idx_v += 1
            self.store_md_generic(metadata, self.ws_v, self.idx_v)
            self.stats.md_types_repartition["vector"] += 1
            self.store_md_vector(metadata, self.ws_v, self.idx_v)
        elif metadata.type == "rasterDataset":
            self.idx_r += 1
            self.store_md_generic(metadata, self.ws_r, self.idx_r)
            self.stats.md_types_repartition["raster"] += 1
            self.store_md_raster(metadata, self.ws_r, self.idx_r)
        elif metadata.type == "service":
            self.idx_s += 1
            self.store_md_generic(metadata, self.ws_s, self.idx_s)
            self.stats.md_types_repartition["service"] += 1
            self.store_md_service(metadata, self.ws_s, self.idx_s)
        elif metadata.type == "resource":
            self.idx_rz += 1
            self.store_md_generic(metadata, self.ws_rz, self.idx_rz)
            self.stats.md_types_repartition["resource"] += 1
            self.store_md_resource(metadata, self.ws_rz, self.idx_rz)
        else:
            logger.error(
                "Type of metadata is not recognized/handled: {}".format(metadata.type)
            )

    def store_md_generic(self, md: Metadata, ws: Worksheet, idx: int):
        """Exports generic metadata attributes into Excel worksheet with some dynamic
        adaptations based on metadata type.

        :param Metadata md: metadata object to export
        :param Worksheet ws: Excel worksheet to store the exported info
        :param int idx: row index in the worksheet
        """
        # pick columns referential table depending on metadata type
        if md.type == "rasterDataset":
            col = self.columns_raster
        elif md.type == "resource":
            col = self.columns_resource
        elif md.type == "service":
            col = self.columns_service
        elif md.type == "vectorDataset":
            col = self.columns_vector
        else:
            raise TypeError("Unknown metadata type: {}".format(md.type))

        logger.debug(
            "Start storing metadata {} ({}) using the matching reference columns for type of {} ...".format(
                md.title_or_name(slugged=1), md._id, md.type
            )
        )

        # -- IDENTIFICATION ------------------------------------------------------------
        if md.title:
            ws["{}{}".format(col.get("title").letter, idx)] = md.title
        if md.name:
            ws["{}{}".format(col.get("name").letter, idx)] = md.name
        if md.abstract:
            ws["{}{}".format(col.get("abstract").letter, idx)] = md.abstract

        # path to source
        try:
            src_path = Path(str(md.path))
        except OSError as e:
            logger.debug(
                "Metadata.path value is not a valid system path. Maybe an URL? Original error: {}".format(
                    e
                )
            )
            urlparse(md.path).scheme != ""

        if isinstance(md.path, Path) and md.type != "service":
            if src_path.is_file():
                link_path = r'=HYPERLINK("{0}","{1}")'.format(
                    src_path.parent, src_path.resolve()
                )
                ws["{}{}".format(col.get("path").letter, idx)] = link_path
                logger.debug("Path reachable: {}".format(src_path))
            else:
                ws["{}{}".format(col.get("path").letter, idx)] = str(src_path.resolve())
                logger.debug(
                    "Path not recognized nor reachable: {}".format(str(src_path))
                )
        elif md.path and md.type == "service":
            link_path = r'=HYPERLINK("{0}","{1}")'.format(md.path, md.path)
            ws["{}{}".format(col.get("path").letter, idx)] = link_path
        else:
            pass

        # -- TAGS ----------------------------------------------------------------------
        keywords = []
        inspire = []
        if md.keywords:
            for k in md.keywords:
                if k.get("_tag").startswith("keyword:is"):
                    keywords.append(k.get("text"))
                elif k.get("_tag").startswith("keyword:in"):
                    inspire.append(k.get("text"))
                else:
                    logger.info("Unknown keyword type: " + k.get("_tag"))
                    continue
            if keywords:
                ws["{}{}".format(col.get("keywords").letter, idx)] = " ;\n".join(
                    sorted(keywords)
                )
            if inspire:
                ws["{}{}".format(col.get("inspireThemes").letter, idx)] = " ;\n".join(
                    sorted(inspire)
                )
        else:
            self.stats.md_empty_fields[md._id].append("keywords")
            logger.info("Vector dataset without any keyword or INSPIRE theme")

        # INSPIRE conformity
        ws["{}{}".format(col.get("inspireConformance").letter, idx)] = (
            "conformity:inspire" in md.tags
        )

        # owner
        ws["{}{}".format(col.get("_creator").letter, idx)] = next(
            v for k, v in md.tags.items() if "owner:" in k
        )

        # -- HISTORY -------------------------------------------------------------------
        if md.collectionContext:
            ws[
                "{}{}".format(col.get("collectionContext").letter, idx)
            ] = md.collectionContext
        if md.collectionMethod:
            ws[
                "{}{}".format(col.get("collectionMethod").letter, idx)
            ] = md.collectionMethod

        # validity
        if md.validFrom:
            ws["{}{}".format(col.get("validFrom").letter, idx)] = utils.hlpr_datetimes(
                md.validFrom
            )
            ws["{}{}".format(col.get("validFrom").letter, idx)].style = "date"

        if md.validTo:
            ws["{}{}".format(col.get("validTo").letter, idx)] = utils.hlpr_datetimes(
                md.validTo
            )
            ws["{}{}".format(col.get("validTo").letter, idx)].style = "date"

        if md.updateFrequency:
            ws[
                "{}{}".format(col.get("updateFrequency").letter, idx)
            ] = md.updateFrequency
        if md.validityComment:
            ws[
                "{}{}".format(col.get("validityComment").letter, idx)
            ] = md.validityComment

        # -- EVENTS --------------------------------------------------------------------
        # data creation date
        if md.created:
            ws["{}{}".format(col.get("created").letter, idx)] = utils.hlpr_datetimes(
                md.created
            )
            ws["{}{}".format(col.get("created").letter, idx)].style = "date"

        # events count
        if md.events:
            ws["{}{}".format(col.get("events").letter, idx)] = len(md.events)

        # data last update
        if md.modified:
            ws["{}{}".format(col.get("modified").letter, idx)] = utils.hlpr_datetimes(
                md.modified
            )
            ws["{}{}".format(col.get("modified").letter, idx)].style = "date"

        # -- TECHNICAL -----------------------------------------------------------------
        # format
        if md.format and md.type in ("rasterDataset", "vectorDataset"):
            format_lbl = next(v for k, v in md.tags.items() if "format:" in k)
            ws["{}{}".format(col.get("format").letter, idx)] = "{0} ({1} - {2})".format(
                format_lbl, md.formatVersion, md.encoding
            )
        elif md.format:
            ws["{}{}".format(col.get("format").letter, idx)] = "{0} {1}".format(
                md.format, md.formatVersion
            )
        else:
            pass

        # SRS
        if isinstance(md.coordinateSystem, dict):
            ws[
                "{}{}".format(col.get("coordinateSystem").letter, idx)
            ] = "{0} ({1})".format(
                md.coordinateSystem.get("name"), md.coordinateSystem.get("code")
            )

        # bounding box (envelope)
        if md.envelope and md.envelope.get("bbox"):
            coords = md.envelope.get("coordinates")
            if md.envelope.get("type") == "Polygon":
                bbox = ",\n".join(
                    format(coord, ".4f") for coord in md.envelope.get("bbox")
                )
            elif md.envelope.get("type") == "Point":
                bbox = "Centroïde : {}{}".format(coords[0], coords[1])
            else:
                bbox = ",\n".join(
                    format(coord, ".4f") for coord in md.envelope.get("bbox")
                )
            ws["{}{}".format(col.get("envelope").letter, idx)] = bbox

        # geometry
        if md.geometry:
            ws["{}{}".format(col.get("geometry").letter, idx)] = md.geometry

        # resolution
        if md.distance:
            ws["{}{}".format(col.get("distance").letter, idx)] = md.distance

        # scale
        if md.scale:
            ws["{}{}".format(col.get("scale").letter, idx)] = md.scale

        # features objects
        if md.features:
            ws["{}{}".format(col.get("features").letter, idx)] = md.features

        # -- QUALITY -------------------------------------------------------------------
        if md.specifications:
            ws["{}{}".format(col.get("specifications").letter, idx)] = " ;\n".join(
                self.fmt.specifications(md.specifications)
            )

        # topology
        if md.topologicalConsistency:
            ws["AC{}".format(idx)] = md.topologicalConsistency

        # -- FEATURE ATTRIBUTES --------------------------------------------------------
        if md.type == "vectorDataset" and isinstance(md.featureAttributes, list):
            fields = md.featureAttributes

            # count
            ws["{}{}".format(col.get("featureAttributesCount").letter, idx)] = len(
                fields
            )
            # alphabetic list
            fields_cct = sorted(
                [
                    "{} ({}) - Type : {} - Descripion : {:.20} [...]".format(
                        field.get("name"),
                        field.get("alias"),
                        field.get("dataType"),
                        # field.get("language"),
                        field.get("description", ""),
                    )
                    for field in fields
                ]
            )
            ws["{}{}".format(col.get("featureAttributes").letter, idx)] = " ;\n".join(
                fields_cct
            )
            # if attributes analisis is activated, append fields dict
            if hasattr(self, "ws_fa"):
                self.fa_all.append(fields)
            else:
                pass

        # -- CGUs ----------------------------------------------------------------------
        if md.conditions:
            ws["{}{}".format(col.get("conditions").letter, idx)] = " ;\n".join(
                self.fmt.conditions(md.conditions)
            )

        # -- LIMITATIONS ---------------------------------------------------------------
        if md.limitations:
            ws["{}{}".format(col.get("limitations").letter, idx)] = " ;\n".join(
                self.fmt.limitations(md.limitations)
            )

        # -- CONTACTS ------------------------------------------------------------------
        if md.contacts:
            contacts = [
                "{0} ({1})".format(
                    contact.get("contact").get("name"),
                    contact.get("contact").get("email"),
                )
                for contact in md.contacts
            ]
            ws["{}{}".format(col.get("contacts").letter, idx)] = " ;\n".join(contacts)

        # -- ACTIONS -------------------------------------------------------------------
        ws["{}{}".format(col.get("hasLinkDownload").letter, idx)] = (
            "action:download" in md.tags
        )
        ws["{}{}".format(col.get("hasLinkView").letter, idx)] = "action:view" in md.tags
        ws["{}{}".format(col.get("hasLinkOther").letter, idx)] = (
            "action:other" in md.tags
        )

        # -- METADATA ------------------------------------------------------------------
        # id
        ws["{}{}".format(col.get("_id").letter, idx)] = md._id

        # creation
        if md._created:
            ws["{}{}".format(col.get("_created").letter, idx)] = utils.hlpr_datetimes(
                md._created
            )
            ws["{}{}".format(col.get("_created").letter, idx)].style = "date"

        # last update
        if md._modified:
            ws["{}{}".format(col.get("_modified").letter, idx)] = utils.hlpr_datetimes(
                md._modified
            )
            ws["{}{}".format(col.get("_modified").letter, idx)].style = "date"

        # edit
        # ws["{}{}".format(col.get("linkEdit").letter, idx)] = md.admin_url(self.url_base_edit) + "identification"
        ws["{}{}".format(col.get("linkEdit").letter, idx)] = utils.get_edit_url(md)
        if self.share is not None:
            link_visu = utils.get_view_url(
                md_id=md._id, share_id=self.share._id, share_token=self.share.urlToken
            )
            ws["{}{}".format(col.get("linkView").letter, idx)] = link_visu

        # lang
        ws["{}{}".format(col.get("language").letter, idx)] = md.language

    def store_md_vector(self, md: Metadata, ws: Worksheet, idx: int):
        """ TO DOCUMENT
        """

        # STYLING
        ws.row_dimensions[idx].height = 35  # line height - see #52
        ws["C{}".format(idx)].style = "wrap"
        ws["F{}".format(idx)].style = "wrap"
        ws["G{}".format(idx)].style = "wrap"
        ws["I{}".format(idx)].style = "wrap"
        ws["J{}".format(idx)].style = "wrap"
        ws["K{}".format(idx)].style = "date"
        ws["L{}".format(idx)].style = "date"
        ws["U{}".format(idx)].style = "wrap"
        ws["AA{}".format(idx)].style = "wrap"
        ws["AB{}".format(idx)].style = "wrap"
        ws["AC{}".format(idx)].style = "wrap"
        ws["AD{}".format(idx)].style = "wrap"
        ws["AE{}".format(idx)].style = "wrap"
        ws["AG{}".format(idx)].style = "wrap"
        ws["AH{}".format(idx)].style = "wrap"

        # LOG
        logger.info(
            "Vector metadata stored: {} ({})".format(
                md.title_or_name(slugged=1), md._id
            )
        )

    def store_md_raster(self, md: Metadata, ws: Worksheet, idx: int):
        """ TO DOCUMENT
        """

        # STYLING
        ws.row_dimensions[idx].height = 35  # line height - see #52
        ws["C{}".format(idx)].style = "wrap"
        ws["F{}".format(idx)].style = "wrap"
        ws["G{}".format(idx)].style = "wrap"
        ws["I{}".format(idx)].style = "wrap"
        ws["J{}".format(idx)].style = "wrap"
        ws["K{}".format(idx)].style = "date"
        ws["L{}".format(idx)].style = "date"
        ws["U{}".format(idx)].style = "wrap"
        ws["X{}".format(idx)].style = "wrap"
        ws["Y{}".format(idx)].style = "wrap"
        ws["Z{}".format(idx)].style = "wrap"
        ws["AA{}".format(idx)].style = "wrap"
        ws["AC{}".format(idx)].style = "wrap"
        ws["AD{}".format(idx)].style = "wrap"

        # LOG
        logger.info("Raster metadata stored: {} ({})".format(md.name, md._id))

        # end of method
        return

    def store_md_service(self, md: Metadata, ws: Worksheet, idx: int):
        """ TO DOCUMENT
        """

        # STYLING
        ws.row_dimensions[idx].height = 35  # line height - see #52
        ws["C{}".format(idx)].style = "wrap"
        ws["F{}".format(idx)].style = "wrap"
        ws["M{}".format(idx)].style = "wrap"
        ws["N{}".format(idx)].style = "wrap"
        ws["O{}".format(idx)].style = "wrap"
        ws["P{}".format(idx)].style = "wrap"
        ws["R{}".format(idx)].style = "wrap"
        ws["S{}".format(idx)].style = "wrap"

        # LOG
        logger.info("Service metadata stored: {} ({})".format(md.name, md._id))

        # end of method
        return

    def store_md_resource(self, md: Metadata, ws: Worksheet, idx: int):
        """ TO DOCUMENT
        """

        # STYLING
        ws.row_dimensions[idx].height = 35  # line height - see #52
        ws["C{}".format(idx)].style = "wrap"
        ws["F{}".format(idx)].style = "wrap"
        ws["M{}".format(idx)].style = "wrap"
        ws["N{}".format(idx)].style = "wrap"
        ws["O{}".format(idx)].style = "wrap"
        ws["P{}".format(idx)].style = "wrap"
        ws["R{}".format(idx)].style = "wrap"
        ws["S{}".format(idx)].style = "wrap"

        # LOG
        logger.info("Resource metadata stored: {} ({})".format(md.name, md._id))

        # end of method
        return

    # ------------ Analisis --------------------------------------------------
    def analisis_attributes(self):
        """Perform feature attributes analisis and write results into the
        dedicatedWworksheet."""
        # local arrays
        fa_names = []
        fa_types = []
        fa_alias = []
        fa_descr = []

        # parsing
        for dico_fa in self.fa_all:
            for fa in dico_fa:
                fa_names.append(fa.get("name"))
                # fa_alias.append(fa.get("alias", "NR"))
                # fa_types.append(fa.get("dataType"))
                # fa_descr.append(fa.get("description", "NR"))
                del fa

        # stats
        frq_names = Counter(fa_names)
        frq_alias = Counter(fa_alias)
        frq_types = Counter(fa_types)
        frq_descr = Counter(fa_descr)

        # write
        ws = self.ws_fa
        for fa in frq_names:
            self.idx_fa += 1
            ws["A{}".format(self.idx_fa)] = fa
            ws["B{}".format(self.idx_fa)] = frq_names.get(fa)

    # ------------ CustomizeWworksheet ----------------------------------------
    def headers_writer(self, ws: Worksheet, columns: ColumnPattern):
        """Writes the headers from a columns ref table to a worksheet.

        Arguments:
            ws {Worksheet} -- worksheet into write headers
            columns {ColumnPattern} -- column table
        """
        # text
        for k, v in columns.items():
            if v.letter is None:
                continue
            ws["{}1".format(v.letter)] = self.tr.get(k, "Missing translation")

        # styling
        for row_cols in ws.iter_cols(min_row=1, max_row=1):
            for cell in row_cols:
                cell.style = "Headline 2"

    def tunning_worksheets(self):
        """Automate"""
        for sheet in self.worksheets:
            # Freezing panes
            c_freezed = sheet["B2"]
            sheet.freeze_panes = c_freezed

            # Print properties
            sheet.print_options.horizontalCentered = True
            sheet.print_options.verticalCentered = True
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

            # Others properties
            wsprops = sheet.sheet_properties
            wsprops.filterMode = True

            # enable filters
            sheet.auto_filter.ref = str("A1:{}{}").format(
                get_column_letter(sheet.max_column), sheet.max_row
            )


# #############################################################################
# ##### Stand alone program ########
# ##################################
if __name__ == "__main__":
    """ Standalone execution and development tests """
    # ------------ Specific imports ----------------
    from dotenv import load_dotenv
    from logging.handlers import RotatingFileHandler
    from os import environ
    import urllib3

    from isogeo_pysdk import Isogeo

    # ------------ Log & debug ----------------
    logger = logging.getLogger()
    logging.captureWarnings(True)
    logger.setLevel(logging.DEBUG)
    # logger.setLevel(logging.INFO)

    log_format = logging.Formatter(
        "%(asctime)s || %(levelname)s "
        "|| %(module)s - %(lineno)d ||"
        " %(funcName)s || %(message)s"
    )

    # debug to the file
    log_file_handler = RotatingFileHandler("dev_debug.log", "a", 3000000, 1)
    log_file_handler.setLevel(logging.DEBUG)
    log_file_handler.setFormatter(log_format)

    # info to the console
    log_console_handler = logging.StreamHandler()
    log_console_handler.setLevel(logging.INFO)
    log_console_handler.setFormatter(log_format)

    logger.addHandler(log_file_handler)
    logger.addHandler(log_console_handler)

    # ------------ Real start ----------------
    # get user ID as environment variables
    load_dotenv("dev.env")

    # ignore warnings related to the QA self-signed cert
    if environ.get("ISOGEO_PLATFORM").lower() == "qa":
        urllib3.disable_warnings()

    # for oAuth2 Backend (Client Credentials Grant) Flow
    isogeo = Isogeo(
        auth_mode="group",
        client_id=environ.get("ISOGEO_API_GROUP_CLIENT_ID"),
        client_secret=environ.get("ISOGEO_API_GROUP_CLIENT_SECRET"),
        auto_refresh_url="{}/oauth/token".format(environ.get("ISOGEO_ID_URL")),
        platform=environ.get("ISOGEO_PLATFORM", "qa"),
    )

    # getting a token
    isogeo.connect()

    # misc
    METADATA_TEST_FIXTURE_UUID = environ.get("ISOGEO_FIXTURES_METADATA_COMPLETE")
    WORKGROUP_TEST_FIXTURE_UUID = environ.get("ISOGEO_WORKGROUP_TEST_UUID")

    search = isogeo.search(
        whole_results=0,
        query="owner:{}".format(WORKGROUP_TEST_FIXTURE_UUID),
        include="all",
    )

    isogeo.close()  # close session

    print(
        "{}/{} metadata ready to be exported.".format(len(search.results), search.total)
    )

    # instanciate th final workbook
    out_workbook = Isogeo2xlsx(
        lang=isogeo.lang, url_base_edit=isogeo.app_url, url_base_view=isogeo.oc_url
    )
    # add needed worksheets
    out_workbook.set_worksheets(auto=search.tags.keys())

    # parse search results
    for md in map(Metadata.clean_attributes, search.results):
        out_workbook.store_metadatas(md)

    # save file
    out_workbook.save("test_isogeo_export_to_xlsx.xlsx")
