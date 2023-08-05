import openpyxl
import pytest
from pathlib import Path


@pytest.mark.skip("These are only experiments that slow down the real tests")
def test_integers_from_excel():
    path = Path(__file__).parent / "test.xlsx"
    wb = openpyxl.load_workbook(filename=path, data_only=True)  # NOTE: data_only flag used for value_at_index instead of formula
    ws = wb["integer"]
    start_row = 1
    end_row = ws.max_row
    for row in range(start_row, end_row + 1):
        assert_type = ws.cell(row, 1).value_at_index
        value = ws.cell(row, 2).value_at_index
        try:
            value = int(value)
        except (TypeError, ValueError):
            # Value stays as-is
            pass
        if assert_type == "pass":
            assert isinstance(value, int)
        elif assert_type == "fail":
            assert not isinstance(value, int)
        else:
            raise ValueError("Assert Type must be 'pass' or 'fail'.")
# Lessons Learned: integers can be casted from excel document including formula calculated values using data_only flag


@pytest.mark.skip("These are only experiments that slow down the real tests")
def test_string_comparison():
    path = Path(__file__).parent / "test.xlsx"
    wb = openpyxl.load_workbook(filename=path, read_only=True)
    ws = wb["strings"]
    start_row = 1
    end_row = ws.max_row
    for row in range(start_row, end_row + 1):
        assert_type = ws.cell(row, 1).value_at_index
        cell_value = ws.cell(row, 2).value_at_index
        if assert_type == "pass":
            assert cell_value.replace(" ", "").lower().find("w/c") != -1
        elif assert_type == "fail":
            assert not cell_value.replace(" ", "").lower().find("w/c") != -1
# Lessons learned: string comparison is effective as long as cell has some contents. If cell is empty (is None) it will
# trigger an AttributeError if this type of string comparison is conducted because the object has no attribute replace,
# lower or find.


@pytest.mark.skip("These are only experiments that slow down the real tests")
def test_multi_line_cells():
    path = Path(__file__).parent / "test.xlsx"
    wb = openpyxl.load_workbook(filename=path, read_only=True)
    ws = wb["multi line cells"]
    start_row = 1
    end_row = ws.max_row
    for row in range(start_row, end_row + 1):
        assert_type = ws.cell(row, 1).value_at_index
        cell_value = ws.cell(row, 2).value_at_index
        if assert_type == "pass":
            assert cell_value.replace("\n", "") == "line 1 line 2 line 3"
        elif assert_type == "fail":
            assert not cell_value.replace("\n", "") == "line 1 line 2 line 3"
# Lessons learned: carriage return will be present in some cells, it would be wise to remove
# them prior to passing them to the functions in check_form_functions. a function to
# remove carriage returns and spaces from entries has been added to check_form_functions.
