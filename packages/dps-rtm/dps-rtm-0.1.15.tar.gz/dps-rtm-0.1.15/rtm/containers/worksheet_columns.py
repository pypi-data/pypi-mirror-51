# --- Standard Library Imports ------------------------------------------------
from collections import namedtuple
from typing import List

# --- Third Party Imports -----------------------------------------------------
import openpyxl

# --- Intra-Package Imports ---------------------------------------------------
from rtm.main.exceptions import RTMValidatorFileError
import rtm.main.context_managers as context


WorksheetColumn = namedtuple("WorksheetColumn", "header body index column")


class WorksheetColumns:

    def __init__(self, worksheet_name):

        # --- Attributes ------------------------------------------------------
        self.max_row = None
        self.cols = None
        self.body_length = 0

        # --- Get Path ----------------------------------------------------
        path = context.path.get()

        # --- Get Workbook ------------------------------------------------
        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)

        # --- Get Worksheet -----------------------------------------------
        ws = None
        for sheetname in wb.sheetnames:
            if sheetname.lower() == worksheet_name.lower():
                ws = wb[sheetname]
        if ws is None:
            raise RTMValidatorFileError(
                f"\nError: Workbook does not contain a '{worksheet_name}' worksheet"
            )
        self.max_row = ws.max_row
        self.body_length = self.max_row - 1

        # --- Convert Worksheet to WorksheetColumn objects ----------------
        ws_data = []
        start_column_num = 1
        for index, col in enumerate(range(start_column_num, ws.max_column + 1)):
            column_header = ws.cell(1, col).value
            column_body = tuple(ws.cell(row, col).value for row in range(2, self.max_row + 1))
            ws_column = WorksheetColumn(
                header=column_header, body=column_body, index=index, column=col
            )
            ws_data.append(ws_column)
        self.cols = ws_data

    def __getitem__(self, index):
        return self.cols[index]

    def __len__(self):
        return len(self.cols)

    def get_first(self, header_name):
        """returns the first worksheet_column that matches the header"""
        matches = get_matching_worksheet_columns(self, header_name)
        if len(matches) > 0:
            return matches[0]
        else:
            return None


def get_matching_worksheet_columns(sequence_worksheet_columns, field_name) -> List[WorksheetColumn]:
    """Called by constructor to get matching WorksheetColumn objects"""
    matching_worksheet_columns = [
        ws_col
        for ws_col in sequence_worksheet_columns
        if ws_col.header.lower() == field_name.lower()
    ]
    return matching_worksheet_columns
